import os
import openpyxl

excel_path = r"d:\AppleICloud\iCloudDrive\2026SummerTrip\assets\summer plan.xlsx"
out_txt = r"d:\AppleICloud\iCloudDrive\2026SummerTrip\scratch\excel_sheet1.txt"

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb['Sheet1']
        with open(out_txt, 'w', encoding='utf-8') as f:
            for r in range(1, min(sheet.max_row + 1, 100)):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                f.write(f"Row {r:02d}: " + "\t|\t".join([str(val) for val in row_vals]) + "\n")
        print(f"Dumped Sheet1 to {out_txt} successfully!")
    except Exception as e:
        print("Error dumping Excel:", e)
else:
    print("Excel file not found")
