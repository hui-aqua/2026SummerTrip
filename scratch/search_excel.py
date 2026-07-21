import os
import openpyxl

excel_path = r"d:\AppleICloud\iCloudDrive\2026SummerTrip\assets\summer plan.xlsx"

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print("Excel Loaded successfully!")
        
        # Let's search for "ICMCF", "presentation", "oral", "poster", "Cheng", "Yang"
        keywords = ["icmcf", "presentation", "oral", "poster", "cheng", "yang", "speech", "talk", "报告", "会议"]
        
        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"\n--- Searching Sheet: {name} ({sheet.max_row} rows, {sheet.max_column} cols) ---")
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                row_str = " ".join([str(val) for val in row_vals if val is not None])
                
                # Check for keywords
                for kw in keywords:
                    if kw in row_str.lower():
                        print(f"Row {r} matched '{kw}': {row_str[:300]}")
                        break
    except Exception as e:
        print("Error searching Excel:", e)
else:
    print("Excel file not found")
