import os
import openpyxl

excel_path = r"d:\AppleICloud\iCloudDrive\2026SummerTrip\assets\summer plan.xlsx"

if os.path.exists(excel_path):
    try:
        # Load workbook (read_only=True to be memory efficient and fast for large files)
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        print("Sheet Names:", wb.sheetnames)
        
        # Search for conference-related sheets
        for sheetname in wb.sheetnames:
            if "berlin" in sheetname.lower() or "icmcf" in sheetname.lower() or "conference" in sheetname.lower() or "agenda" in sheetname.lower():
                print("Found interesting sheet:", sheetname)
    except Exception as e:
        print("Error reading Excel file:", e)
else:
    print("Excel file not found")
